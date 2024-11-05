# -*- coding: utf-8 -*-
"""
Created on Fri Apr  5 17:41:55 2024
@author: David Chaparro
"""
#librerias
import pandas as pd
from openpyxl.styles import Border,Side
from openpyxl import load_workbook
import os
#funciones
def num_req(ws):
    i = ws['F7'].value
    ws['F7'] = i + 1
    
def insert_fila(ws):
    ws.insert_rows(15)
    bordesnegros = Border(left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
    fila = ws["A15":"F15"]
    for celda in fila[0]:    
        celda.border = bordesnegros
def guardar_datos(df,ws):
    for i in range(len(df)):
        insert_fila(ws)
        ws['A15'] = df.iloc[i,0]
        ws['B15'] = df.iloc[i,1]
        ws['C15'] = df.iloc[i,2]
        ws['D15'] = df.iloc[i,3]
def imprimir(requisicion):
    comand = f'start excel {requisicion}'
    os.system(comand)
#lectura de hoja excel
df = pd.read_csv(r"\\VENTAS\Compartidos\SERVICIOS\BASEDEDATOS\requisicion.csv")
rq = load_workbook(r"\\VENTAS\Compartidos\SERVICIOS\BASEDEDATOS\Requisiciones_.xlsx")
ws = rq['Hoja1']
requisicion = r"\\VENTAS\Compartidos\SERVICIOS\BASEDEDATOS\Requisiciones_print.xlsx"
#cuerpo
num_req(ws)
guardar_datos(df, ws) 
#guardar documento
rq.save('Requisiciones_print.xlsx')
#abrir excel
imprimir(requisicion)